import cv2
import openpyxl
from FaceDetection.face_detection import face
from keras.models import load_model
import numpy as np
import os
import csv
from embedding import emb
from MongoDB.retrieve_pymongo_data import database
import face_recognition
from datetime import datetime
from datetime import date
now=date.today()

import warnings
warnings.filterwarnings("ignore")




def Recognition(subject):
    label=None
    #print("hello world")

    people=sorted(os.listdir('people'))


    lecture=subject


    person=None
    ###########################   code has been change    ######################
    
                   # #       print(os.path.join(root, name))

    video_capture = cv2.VideoCapture(0)


    rootdir = '.\people'

    images=[]####### first load image in jpg formate     ##########
    binary_image=[] #############  second load image in binary image formate    ##########3
    encodings_list=[]##############   third start encoding the image in biomatric form     ################
    known_name_list=[]################    fourth   listing names to compare with biometric formn      ##################
    for subdir, dirs, files in os.walk(rootdir):
        for file in files:
            #print (os.path.join(subdir, file))
            images.append(os.path.join(subdir, file))
            known_name_list.append(os.path.splitext(file)[0])
    #print(known_name_list)

    for item in images:
        binary_image.append(face_recognition.load_image_file(item))
    
    
    for element in binary_image:
        encodings_list.append(face_recognition.face_encodings(element)[0])




    # Initialize some variables
    face_locations = []
    face_encodings = []
    face_names = []
    process_this_frame = True
    ##############################################     attendance function          #####################
    def markattendance(name,frames):
        wb=openpyxl.load_workbook('Attendance.xlsx')

        #print(wb.get_sheet_names())
        sheet=wb.active
        #cell=sheet.cell(row=1,column=1)
        variable=[]
        josh='b6'
        max_rows=sheet.max_row
        for i in range(1, max_rows+1):
            cells=sheet.cell(row=i,column=2)
            #print(cells.value)
            if name == cells.value:
                variable=cells
                variable=str(variable).split('.')
                for i in variable:
                    josh=i       
                #variable=variable.split(',')
                #print(josh)                        
                josh=int(josh[1])
                # print(josh)
                #present=("P","None")
                max_columns=sheet.max_column
                #for i in range(3,int(now.day)+3):
                cells=sheet.cell(row=josh,column=int(now.day)+2)
                #print(josh)                              #### josh = 1,2,3,4,5,6, ....row number from name cell address
                if 'P' == cells.value:
                    # for i in range(100):
                    #     print(i)
                    #     print(cells.value)
                    print("already present")
                else:
                    cells.value='P'
                    # print(cells.value)
                    # print(cells)
                    # print(sheet)
                    wb.save('Attendance.xlsx')
                    #i=max_columns+1
            # else:
            #     cells.value=name
            #     wb.save('Attendance.xlsx')
            #     now=datetime.now()
            #     dtstring=now.strftime('%H:%M:%S')
            #     f.writelines(f'\n{name},{dtstring},present')


    while True: 
        # Grab a single frame of video
        ret, frame = video_capture.read()

        # Resize frame of video to 1/4 size for faster face recognition processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

        # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
        rgb_small_frame = small_frame[:, :, ::-1]

        # Only process every other frame of video to save time
        if process_this_frame:
            # Find all the faces and face encodings in the current frame of video
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            face_names = []
            for face_encoding in face_encodings:
                # See if the face is a match for the known face(s)
                matches = face_recognition.compare_faces(encodings_list, face_encoding)
                name = "Unknown"

                # # If a match was found in encodings_list, just use the first one.
                # if True in matches:
                #     first_match_index = matches.index(True)
                #     name = known_name_list[first_match_index]

                # Or instead, use the known face with the smallest distance to the new face
                face_distances = face_recognition.face_distance(encodings_list, face_encoding)
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_name_list[best_match_index]
                    markattendance(name,frame)

                face_names.append(name)

        process_this_frame = not process_this_frame
        ##############################################################################


        ##########################################################

        # Display the results
        for (top, right, bottom, left), name in zip(face_locations, face_names):
            # Scale back up face locations since the frame we detected in was scaled to 1/4 size
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            # Draw a box around the face
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

            # Draw a label with a name below the face
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

        # Display the resulting image
        cv2.imshow('Say Cheese and Press "Q" to Quite', frame)

        # Hit 'q' on the keyboard to quit!
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release handle to the webcam
    video_capture.release()
    cv2.destroyAllWindows()




